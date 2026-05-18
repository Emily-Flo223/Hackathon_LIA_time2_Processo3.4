"""
nodes/emit.py — Nó de Saída e Entrega (Nó 8) do agente PROECE 3.4

Responsabilidade:
  - Cria um diretório físico de saídas usando o nome do próprio relatório.
  - Salva o parecer opinativo em formato Markdown (.md).
  - Salva o parecer estruturado analítico em formato JSON (.json).
  - Salva a minuta de e-mail de notificação regulamentar em formato texto (.txt).
  - Gera/Atualiza a planilha geral consolidada de divergências e estatísticas (.xlsx).
"""

import json
import logging
import openpyxl
from pathlib import Path
from typing import Optional  
from state import AgenteState

logger = logging.getLogger(__name__)

# Definição do ROOT_DIR para localizar a raiz do projeto de forma dinâmica
ROOT_DIR = Path(__file__).parent.parent

def emit_artifacts(state: AgenteState, output_base_dir: Optional[Path] = None) -> dict:
    """
    Nó 8 — emit_artifacts
    Lê os textos e dados gerados e realiza a gravação física no disco rígido.
    """
    state.log.append("[emit] A iniciar a gravação física dos artefactos da auditoria...")

    try:
        nome_relatorio_original = Path(state.arquivo_path).stem if state.arquivo_path else "relatorio_desconhecido"

        base_dir = Path(output_base_dir) if output_base_dir else ROOT_DIR / "output"
        project_output_dir = base_dir / nome_relatorio_original
        project_output_dir.mkdir(parents=True, exist_ok=True)

        id_nome_arquivo = nome_relatorio_original

        # 1. Salva o Parecer Técnico em Markdown (.md)
        parecer_path = project_output_dir / f"parecer_auditoria_{id_nome_arquivo}.md"
        with open(parecer_path, "w", encoding="utf-8") as f:
            f.write(state.parecer_md)
        state.log.append(f"[emit] Parecer Markdown salvo em: {parecer_path.name}")

        # CORREÇÃO CIRÚRGICA: Remove duplicados mantendo a ordem original da lista
        divergencias_limpas = list(dict.fromkeys(state.divergencias)) if state.divergencias else []

        # 2. Prepara e salva o Parecer Estruturado em JSON (.json)
        state.parecer_json = {
            "metadata": {
                "execution_id": state.execution_id,
                "id_relatorio": state.id_relatorio if state.id_relatorio else id_nome_arquivo,
                "protocolo_projeto": state.protocolo_projeto if state.protocolo_projeto else id_nome_arquivo,
                "tipo_relatorio": state.tipo_relatorio if state.tipo_relatorio else "Não identificado",
                "titulo_projeto": state.titulo_projeto,
                "coordenador": state.coordenador,
                "decisao_final": state.decisao
            },
            "auditoria_completude": {
                "passou": state.check_completude.passou if state.check_completude else True,
                "motivo": state.check_completude.motivo if state.check_completude else "",
                "evidencia": state.check_completude.evidencia if state.check_completude else ""
            },
            "auditoria_financeira": {
                "passou": state.check_financeiro.passou if state.check_financeiro else True,
                "motivo": state.check_financeiro.motivo if state.check_financeiro else "",
                "evidencia": state.check_financeiro.evidencia if state.check_financeiro else ""
            },
            "auditoria_horas": {
                "passou": state.check_horas.passou if state.check_horas else True,
                "motivo": state.check_horas.motivo if state.check_horas else "",
                "evidencia": state.check_horas.evidencia if state.check_horas else ""
            },
            "lista_divergencias_acumuladas": divergencias_limpas
        }

        json_path = project_output_dir / f"dados_auditoria_{id_nome_arquivo}.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(state.parecer_json, f, ensure_ascii=False, indent=4)
        state.log.append(f"[emit] Relatório estruturado JSON salvo em: {json_path.name}")

        # 3. Salva a minuta de e-mail (.txt)
        email_path = project_output_dir / f"minuta_email_{id_nome_arquivo}.txt"
        with open(email_path, "w", encoding="utf-8") as f:
            f.write(state.email_txt)
        state.log.append(f"[emit] Minuta do e-mail regulamentar salva em: {email_path.name}")

        # 4. Gera ou Atualiza a Planilha Geral Consolidada (.xlsx)
        try:
            consolidado_path = base_dir / "consolidado_auditoria.xlsx"
            
            if consolidado_path.exists():
                wb = openpyxl.load_workbook(consolidado_path)
                if "Divergencias" not in wb.sheetnames:
                    wb.create_sheet("Divergencias")
                if "Estatisticas" not in wb.sheetnames:
                    wb.create_sheet("Estatisticas")
                ws_div = wb["Divergencias"]
                ws_stats = wb["Estatisticas"]
            else:
                wb = openpyxl.Workbook()
                ws_div = wb.active
                ws_div.title = "Divergencias"
                ws_div.append(["Protocolo", "Decisão", "Qtd Erros", "Detalhe Divergências"])
                ws_stats = wb.create_sheet("Estatisticas")

            # Verifica se o relatório já existe na planilha para sobrescrever (Evita duplicados)
            linha_existente = None
            for row_idx in range(2, ws_div.max_row + 1):
                if str(ws_div.cell(row=row_idx, column=1).value) == str(id_nome_arquivo):
                    linha_existente = row_idx
                    break

            dados_insercao = [
                id_nome_arquivo,
                state.decisao,
                len(divergencias_limpas),
                " | ".join(divergencias_limpas) if divergencias_limpas else "Nenhuma"
            ]

            if linha_existente:
                # Sobrescreve os dados na linha existente
                for col_idx, valor in enumerate(dados_insercao, start=1):
                    ws_div.cell(row=linha_existente, column=col_idx, value=valor)
            else:
                # Adiciona uma nova linha se não existir
                ws_div.append(dados_insercao)

            # Limpa e recalcula as estatísticas
            ws_stats.delete_rows(1, ws_stats.max_row)
            
            total_relatorios = 0
            total_erros = 0
            count_decisoes = {"APROVAR": 0, "DEVOLVER PARA AJUSTES": 0, "DEVOLVER PARA REELABORAÇÃO": 0}
            count_tipos = {}

            for row in ws_div.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                total_relatorios += 1
                
                dec = row[1]
                if dec in count_decisoes:
                    count_decisoes[dec] += 1
                else:
                    count_decisoes[dec] = 1
                    
                qtd = row[2] or 0
                total_erros += qtd
                
                divs = row[3]
                if divs and divs != "Nenhuma":
                    for div in divs.split(" | "):
                        tipo = div.split(":")[0].strip() if ":" in div else "Outro"
                        count_tipos[tipo] = count_tipos.get(tipo, 0) + 1

            ws_stats.append(["Métrica", "Quantidade", "Frequência (%)"])
            ws_stats.append(["Total de Relatórios Analisados", total_relatorios, "100%"])
            ws_stats.append(["", "", ""])
            
            ws_stats.append(["--- DECISÕES GERAIS ---", "", ""])
            for d_name in ["APROVAR", "DEVOLVER PARA AJUSTES", "DEVOLVER PARA REELABORAÇÃO"]:
                c = count_decisoes.get(d_name, 0)
                pct = f"{(c / total_relatorios) * 100:.1f}%" if total_relatorios > 0 else "0%"
                nome_exibicao = "Acertos (APROVAR)" if d_name == "APROVAR" else d_name
                ws_stats.append([nome_exibicao, c, pct])
            
            ws_stats.append(["", "", ""])
            ws_stats.append(["--- TIPOS DE ERROS ENCONTRADOS ---", "", ""])
            ws_stats.append(["Quantidade Total de Erros", total_erros, "100%"])
            for t, c in count_tipos.items():
                pct = f"{(c / total_erros) * 100:.1f}%" if total_erros > 0 else "0%"
                ws_stats.append([f"Erro tipo: {t}", c, pct])

            for cell in ws_stats[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                
            wb.save(consolidado_path)
            state.log.append(f"[emit] Planilha consolidada atualizada em: {consolidado_path.name}")

        except Exception as ex_excel:
            state.log.append(f"[emit] AVISO: Falha ao atualizar planilha consolidada: {str(ex_excel)}")
            print(f"⚠️ Aviso no nó emit_artifacts (Excel): {str(ex_excel)}")

        state.log.append(f"[emit] OK — Todos os artefactos foram persistidos com sucesso na pasta de saídas.")

    except Exception as e:
        state.log.append(f"[emit] ERRO crítico ao gravar artefactos em disco: {str(e)}")
        print(f"❌ Erro no nó emit_artifacts: {str(e)}")

    return {"parecer_json": getattr(state, "parecer_json", {})}